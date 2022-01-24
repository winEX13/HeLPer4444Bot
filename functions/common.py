import functions.database as db
# import database as db
# import functions.weather as weath
import pickle
import os
from fuzzywuzzy import fuzz

import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords, wordnet
from nltk.probability import FreqDist

from wiki_ru_wordnet import WikiWordnet as wikiWordNet

from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans

from langdetect import detect
import re

import random
import sys
import pymorphy2
import enchant
from enchant.checker import SpellChecker
from enchant.tokenize import EmailFilter, URLFilter
import time
import datetime

import functions.imgDetect as img2txt
import functions.audDetect as aud2txt
import functions.pythonChecker as pyCheck
import functions.network.network as net

from PIL import Image
import aiogram.utils.markdown as fmt
from collections import Counter as count
import openpyxl
from googlesearch import search
import wikipedia as wiki
import bs4 as bs
import urllib.request
from urllib.parse import quote
import requests
import imghdr
from wordcloud import WordCloud as wc, STOPWORDS

import matplotlib.pyplot as plt
import numpy as np

# import spacy
# import textacy.extract

# nltk.download('omw-1.4')

dic = enchant.Dict('ru_RU')
checker = SpellChecker('ru_RU', filters=[EmailFilter])

# nlp = spacy.load('ru_core_news_lg')

def load_obj(name):
    with open(name + '.pkl', 'rb') as f:
        return pickle.load(f)

def save_obj(obj, name):
    with open(name + '.pkl', 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def textFmtH(text, fmrt):
    switcher = { 
        'курсив': 'hitalic',
        'жирный': 'hbold',
        'подчеркнутый': 'hunderline',
        'зачеркнутый': 'hstrikethrough',
        'моно': 'hcode',
        'моно блок': 'hpre',
        'ссылка': 'hlink',
        'невидимая ссылка': 'hide_link',
        'цитата': 'quote_html',
        'обычный': 'text'
    }
    
    if switcher.get(fmrt, 'Invalid') == 'Invalid':
        exit('Error: Formats: курсив, жирный, подчеркнутый, зачеркнутый, моно, моно блок, ссылка, невидимая ссылка, цитата, обычный')
    elif switcher.get(fmrt, 'Invalid') == 'hlink':
        print(text[:text.find(' ')], text[text.find(' '):])
        return (getattr(fmt, switcher.get(fmrt, 'Invalid'))(text[:text.find(' ')], text[text.find(' ') + 1:]))
    return (getattr(fmt, switcher.get(fmrt, 'Invalid'))(text))

def keyboardCreate(keyboard, row, column, lst):
    buttons_row = []
    buttons = []
    # for b in lst:
    for c in range(0, column * row, row):
        for r in range(row):
            try:
                buttons_row.append(lst[c+r])
            except IndexError:
                buttons_row.append('')
        buttons.append(buttons_row)
        buttons_row = []
    for row in buttons:
    	if len(row) > 0:
        	keyboard.add(*row)
    return(keyboard)

def inlineKeyboardCreate(keyboard, row, column, lst):
    buttons_row = []
    buttons = []
    # for b in lst:
    for c in range(0, column * row, row):
        for r in range(row):
            try:
                buttons_row.append(types.InlineKeyboardButton(text=str(lst[c+r][0]), callback_data=str(lst[c+r][1])))
            except IndexError:
                pass
                # buttons_row.append('')
        buttons.append(buttons_row)
        buttons_row = []
    for row in buttons:
        if len(row) > 0:
            keyboard.add(*row)
    return(keyboard)

def text_normalization(text):
    text = str(text).lower()
    spl_char_text = re.sub(r'[^ a-zA-Zа-яА-ЯёЁ]', '', text)
    # stopwords = nltk.corpus.stopwords.words('english')
    stopwords = nltk.corpus.stopwords.words('russian')
    stopwords.extend(['что', 'это', 'так', 'вот', 'быть', 'как', 'в', 'к', 'на'])
    # stopwords = []
    tokens = nltk.word_tokenize(spl_char_text)
    stopwords_num = []
    for i, t in enumerate(tokens):
        if t in stopwords:
            stopwords_num.append(i)
    for i in stopwords_num:
        tokens[i] = ''
    while True:
        try:
            tokens.remove('')
        except ValueError:
            break
    lema = nltk.stem.WordNetLemmatizer()
    stemmer = nltk.stem.PorterStemmer()
    tags_list = nltk.pos_tag(tokens, tagset=None)
    lema_words=[]
    for token, pos_token in tags_list:
        if pos_token.startswith('V'):
            pos_val = 'v'
        if pos_token.startswith('J'):
            pos_val = 'a'
        if pos_token.startswith('R'):
            pos_val = 'r'
        else:
            pos_val = 'n'
        lema_token = lema.lemmatize(token, pos_val)
        lema_words.append(lema_token)
    stem_words = [stemmer.stem(token) for token in tokens]
    for i, w in enumerate(lema_words):
        if not dic.check(w):
            try:
                lema_words.pop(i)
                lema_words.insert(i, dic.suggest(w)[0])
            except IndexError:
                pass
    # return(' '.join(stem_words))
    return (' '.join(lema_words))

def text_normalization_lite(text):
    text = str(text).lower()
    spl_char_text = re.sub(r'[^ a-zA-Zа-яА-ЯёЁ]', '', text)
    # stopwords = nltk.corpus.stopwords.words('english')
    stopwords = nltk.corpus.stopwords.words('russian')
    stopwords.extend(['что', 'это', 'так', 'вот', 'быть', 'как', 'в', 'к', 'на'])
    # stopwords = []
    tokens = nltk.word_tokenize(spl_char_text)
    stopwords_num = []
    for i, t in enumerate(tokens):
        if t in stopwords:
            stopwords_num.append(i)
    for i in stopwords_num:
        tokens[i] = ''
    while True:
        try:
            tokens.remove('')
        except ValueError:
            break
    morph = pymorphy2.MorphAnalyzer()
    norm_words = []
    for i in tokens:
        norm_words.append(morph.parse(i)[0].normal_form)
    return (' '.join(norm_words))


# def text_normalization(text):
#     text = str(text).lower()
#     spl_char_text = re.sub(r'[^ a-zA-Zа-яА-ЯёЁ]', '', text)
#     # stopwords = nltk.corpus.stopwords.words('english')
#     stopwords = nltk.corpus.stopwords.words('russian')
#     stopwords.extend(['что', 'это', 'так', 'вот', 'быть', 'как', 'в', 'к', 'на'])
#     # stopwords = []
#     tokens = nltk.word_tokenize(spl_char_text)
#     stopwords_num = []
#     for i, t in enumerate(tokens):
#         if t in stopwords:
#             stopwords_num.append(i)
#     for i in stopwords_num:
#         tokens[i] = ''
#     while True:
#         try:
#             tokens.remove('')
#         except ValueError:
#             break
#     morph = pymorphy2.MorphAnalyzer()
#     lema_words = []
#     tags_list = []
#     for i in tokens:
#         lema_words.append(morph.parse(i)[0].normal_form)
#         tags_list.append(morph.parse(i)[0].tag.POS)
#     return (' '.join(lema_words))
#     # return(lema_words, tags_list)

def cosine_sim_vectors(vec1, vec2):
    vec1 = vec1.reshape(1, -1)
    vec2 = vec2.reshape(1, -1)
    return(cosine_similarity(vec1, vec2)[0][0])

def topic_adder(text):
    text = text_normalization(text)
    try:
        if detect(text) == 'ru':
            tokens = text.split(' ')
            morph = pymorphy2.MorphAnalyzer()
            norm_words = []
            double_words = []
            for i in tokens:
                if 'NOUN' in morph.parse(i)[0].tag:
                    norm_words.append(morph.parse(i)[0].normal_form)
        elif detect(text) == 'en':
            tokens = text.split(' ')
            norm_words = []
            tags_list = nltk.pos_tag(tokens, tagset=None)
            for token, pos_token in tags_list:
                print(token, pos_token)
                if pos_token == 'NN':
                    norm_words.append(token)
        else:
            return('Недостаточно ключевых слов для создания новой темы.')
    except:
        return('Недостаточно ключевых слов для создания новой темы.')
    if len(norm_words) < 2: #Важно изменить на 2!!!
        return('Недостаточно ключевых слов для создания новой темы.')
    for i in norm_words:
        while norm_words.count(i) > 1:
            norm_words.remove(i)
    norm_words.sort()
    norm_words = ('.').join(norm_words[:3])
    if len(db.sqlFind('dict', 'id', f'topic == "{norm_words}"')) == 0:
        id_ = int(db.sqlFind('dict', 'id', f'id != ""')[-1][0])
        if len(db.sqlFind('dict', 'id', f'id == "{id_}"')) != 0:
            id_ += 1
        # print(db.sqlFind('dict', 'id', f'id != ""'))
        db.sqlAdd('dict', [id_, norm_words, text, 'Этот вопрос уже добавлен в словарь и ждет своего ответа.'], 4)
        return(f'В тему "{norm_words}" на позицию {id_} было добавлено: {text}')
    else:
        return(f'Тема "{norm_words}" уже существует')

def dialoge(text, check_per):
    print ('start')
    start_time = time.time()
    db.sqlConnect('bot_db')
    questions = []
    questions_num = []
    for q in db.sqlFind('dict', 'questions', f'id != ""'):
        # questions.append((' ').join(q[0].split(' | ')) + ' ' + text_normalization(text))
        questions.append((' ').join(q[0].split(' | ')))

    for i, q in enumerate(questions):
        vectorizer = CountVectorizer().fit_transform([q, text_normalization(text)])
        vectors = vectorizer.toarray()
        csim = cosine_sim_vectors(vectors[0], vectors[1])
        questions_num.append([i, csim, q])
    max_i = max(questions_num, key=lambda i : i[1])[0]
    answer = db.sqlFind('dict', 'topic, answers', f'id = "{max_i}"')
    # check_per = 0
    if int(questions_num[max_i][1]*100) <= int(check_per):
        print(topic_adder(text))
        db.sqlDisconnect()
        print(time.time() - start_time, ' - ', str(int(questions_num[max_i][1]*100)) + '%')
        # return('Извините, я вас не понимаю.')

    # for i, q in enumerate(questions):
    #     questions_num.append([i, fuzz.partial_ratio(text_normalization(text), text_normalization(q)), text_normalization(q), q])
    # max_i = max(questions_num, key=lambda i : i[1])[0]
    # answer = db.sqlFind('dict', 'topic, answers', f'id = "{max_i}"')
    
    # return(f'{answer[0]}: {random.choice(answer[1].split(' | '))}')
    db.sqlDisconnect()
    print(time.time() - start_time, ' - ', str(int(questions_num[max_i][1]*100)) + '%')
    # return(str(questions_num[max_i][0])+ ' | ' + str(int(questions_num[max_i][1]*100)) + '% | ' + answer[0][0] + ': ' + random.choice(answer[0][1].split(' | ')))
    return(net.netAnswer(text_normalization(text)))

def img2dialoge():
    answer = dialoge(img2txt.img2txt(), 95)
    # os.remove('functions/objects.jpg')
    return(answer)

def aud2dialoge():
    answer = dialoge(aud2txt.aud2txt(), 0)
    # os.remove('functions/objects.jpg')
    return(answer)

def reborn():
    questions = []
    answers = []

    for i in range(1, 577):
        questions.append(openpyxl.load_workbook(filename = 'train_data.xlsx')['Sheet1'].cell(i, 1).value)
        # print(openpyxl.load_workbook(filename = 'train_data.xlsx')['Sheet1'].cell(i, 1).value)
    for i in range(1, 577):
        answers.append(openpyxl.load_workbook(filename = 'train_data.xlsx')['Sheet1'].cell(i, 2).value)

    # for i in range(len(questions)):
    #     print(f'{i}: {questions[i]}\n{dialoge(questions[i], 95)}')

    for i in range(len(questions)):
        # s = 'Как написать объявление константы?'
        d = dialoge(questions[i], 0)
        db.sqlConnect('bot_db')
        try:
            answer = db.sqlFind('dict', 'questions, answers', f'id = "{d[:d.find(" | ")]}"')
            question = answer[0][0].split(' | ')
            answer = answer[0][1].split(' | ')
        except:
            break

        if not (' ').join(questions[i].split()) in question:
            question.append((' ').join(questions[i].split()))
        if not (' ').join(answers[i].split()) in answer:
            answer.append((' ').join(answers[i].split()))
        try:
            answer.remove((' ').join('Этот вопрос уже добавлен в словарь и ждет своего ответа.'.split()))
        except:
            pass
        # answer.remove('k e k  ')

        question = (' | ').join(question)
        answer = (' | ').join(answer)
        db.sqlUpd('dict', f'questions = "{question}"', f'id = "{d[:d.find(" | ")]}"')
        db.sqlUpd('dict', f'answers = "{answer}"', f'id = "{d[:d.find(" | ")]}"')
        # question = db.sqlFind('dict', 'questions, answers', f'id = "{d[:d.find(" | ")]}"')
        answer = db.sqlFind('dict', 'questions, answers', f'id = "{d[:d.find(" | ")]}"')
        print(answer)
        db.sqlDisconnect()

def search_(query):
    folder = 'functions/processing/data/'
    # folder = 'processing/data/'
    answer = []
    answer_num = []
    answers = []
    gif = []
    jpeg = []
    png = []
    webp = []
    article_text = []

    start_time = time.time()

    wiki.set_lang('ru')
    try:
        wiki_text = wiki.summary(text_normalization_lite(query), sentences=5)
        wiki_images = wiki.page(text_normalization_lite(query)).images
    except:
        wiki_text = ''

    for i in search(str(query), tld='ru', lang='ru', num=10, start=0, stop=10, pause=2.0):
        # print(bs.BeautifulSoup(requests.get(i).content, 'html.parser').find_all('img'))
        
        try:
            for a in bs.BeautifulSoup(requests.get(i).content, 'html.parser').find_all('img'):
                if isinstance(a.get('src'), str):
                    if '.gif' in a.get('src') and 'http' in a.get('src'):
                        gif.append(a.get('src')[a.get('src').find('http'):a.get('src').rfind('.gif')+4])
                    elif '.jpg' in a.get('src') and 'http' in a.get('src'):
                        jpeg.append(a.get('src')[a.get('src').find('http'):a.get('src').rfind('.jpg')+4])
                    elif '.png' in a.get('src') and 'http' in a.get('src'):
                        png.append(a.get('src')[a.get('src').find('http'):a.get('src').rfind('.png')+4])
                    elif '.webp' in a.get('src') and 'http' in a.get('src'):
                        webp.append(a.get('src')[a.get('src').find('http'):a.get('src').rfind('.webp')+5])
            
            text_worker = ''
            for a in bs.BeautifulSoup(requests.get(i).content, 'html.parser').find_all('p'):
                text_worker += a.text
            article_text.append(re.sub(r'[\n\xa0]', '', text_worker))
        except:
            pass

    # print(sent_tokenize(article_text))
    # answer = sent_tokenize(article_text)

    # print (('\n').join(gif), ('\n').join(jpeg), ('\n').join(png), ('\n').join(webp))

    # for i in range(10):
    #     for i, a in enumerate(answer):
    #         vectorizer = CountVectorizer().fit_transform([a, str(query)])
    #         vectors = vectorizer.toarray()
    #         csim = cosine_sim_vectors(vectors[0], vectors[1])
    #         answer_num.append([i, csim, a])

    #     max_i = max(answer_num, key=lambda i : i[1])[0]
    #     # answers.append(re.sub(r'[^ a-zA-Zа-яА-ЯЁё]', '', answer_num[max_i][2]).capitalize()+'.')
    #     answers.append(answer_num[max_i][2])
    #     answer_num.pop(max_i)

    words = word_tokenize(text_normalization_lite(article_text))
    # print(text_normalization(query).split(' '))
    hypernyms = []
    hyponyms = []
    for i in text_normalization_lite(query).split():
        try:
            for j in wikiWordNet().get_synsets(i)[0].get_words():
                while j.lemma() in words:
                    words.remove(j.lemma())
        except:
            try:
                for j in wordnet.synsets(i)[0].lemmas():
                    while j.name() in words:
                        words.remove(j.name())
            except:
                pass
        try:
            for l in wikiWordNet().get_hypernyms(wikiWordNet().get_synsets(i)[0]):
                for j in l.get_words():
                    hypernyms.append(j.lemma())
            for l in wikiWordNet().get_hyponyms(wikiWordNet().get_synsets(i)[0]):
                for j in l.get_words():
                    hyponyms.append(j.lemma())
        except:
            try:
                for l in wordnet.synsets(i)[0].hypernyms():
                    for j in l.lemmas():
                        hypernyms.append(j.name())
                for l in wordnet.synsets(i)[0].hyponyms():
                    for j in l.lemmas():
                        hyponyms.append(j.name())
            except:
                pass

    cloud_mask = np.array(Image.open(folder[0:folder.find('processing/')+11] + 'cloud_mask.bmp'))

    if len(hypernyms + hyponyms) > 0:
        wc(background_color='white', 
            max_words=2000, 
            mask=cloud_mask, 
            stopwords=set(STOPWORDS), 
            contour_width=3, 
            contour_color='steelblue').generate((' ').join(hypernyms + hyponyms)).to_file(folder[0:folder.find('processing/')+11] + 'wc1.png')
    else:
        wc(background_color='white', 
            max_words=2000, 
            mask=cloud_mask, 
            stopwords=set(STOPWORDS), 
            contour_width=3, 
            contour_color='steelblue').generate('ПУСТО').to_file(folder[0:folder.find('processing/')+11] + 'wc1.png')

    if len(words) > 0:
        wc(background_color='white', 
            max_words=2000, 
            mask=cloud_mask, 
            stopwords=set(STOPWORDS), 
            contour_width=3, 
            contour_color='steelblue').generate((' ').join(words)).to_file(folder[0:folder.find('processing/')+11] + 'wc2.png')
    else:
        wc(background_color='white', 
            max_words=2000, 
            mask=cloud_mask, 
            stopwords=set(STOPWORDS), 
            contour_width=3, 
            contour_color='steelblue').generate('ПУСТО').to_file(folder[0:folder.find('processing/')+11] + 'wc1.png')

    # fdist = FreqDist(re.sub(r'[^ a-zA-Zа-яА-ЯёЁ.,;:?!-{}()\|/_+=%^&*$#@<]', '', article_text).split())

    # fdist = FreqDist(sent_tokenize(re.sub(r'[\n\xa0]', '', (' ').join(article_text))))
    # sentences = fdist.most_common(100)

    sentences = sent_tokenize(re.sub(r'[\n\xa0]', '', (' ').join(article_text)))
    # print(sentences)

    # for i, s in enumerate(sentences):
    #     if s[1] > 1:
    #         vectorizer = CountVectorizer().fit_transform([s[0], str(text_normalization_lite(query))])
    #         vectors = vectorizer.toarray()
    #         csim = cosine_sim_vectors(vectors[0], vectors[1])
    #         answer_num.append([i, csim, s[0]])
    #     else:
    #         break

    # answer_num = sorted(answer_num, key=lambda i : i[1], reverse=True)
    
    # for i in answer_num:
    #     if i[1] > 0:
    #         answers.append(i[2])
    #     else:
    #         break

    # word_count = CountVectorizer().fit_transform(article_text)
    tfidf_vectorizer = TfidfVectorizer(use_idf=True)
    tfidf_vectorizer_vectors = tfidf_vectorizer.fit_transform(article_text)
    names = tfidf_vectorizer.get_feature_names_out()
    vectors = np.asarray(tfidf_vectorizer_vectors[1].T.todense()).tolist()

    for i in range(len(names)):
        answer_num.append([names[i], vectors[i][0]])

    answer_num = sorted(answer_num, key=lambda i : i[1], reverse=True)

    for i in answer_num:
        if i[1] > 0:
            answers.append(i[0])
        else:
            break

    answers = text_normalization(answers).split()
    print (answers)

    answer_num = []
    for i, s in enumerate(sentences):
        vectorizer = CountVectorizer().fit_transform([s, (' ').join(answers[:])])
        vectors = vectorizer.toarray()
        csim = cosine_sim_vectors(vectors[0], vectors[1])
        if csim > 0:
            answer_num.append([i, csim, s])

    # print(answer_num)
    answer_num = sorted(answer_num, key=lambda i : i[1], reverse=True)

    answers = []
    for i in answer_num:
        answers.append(i[2])

    for i in answers:
        while answers.count(i) > 1:
            answers.remove(i)

    # print(tfidf_vectorizer_vectors[1], tfidf_vectorizer.get_feature_names())
    # first_vector_tfidfvectorizer = tfidf_vectorizer_vectors[1]


    for i in os.listdir(folder):
        os.remove(folder+i)

    for i in range(len(wiki_images+jpeg+png+webp)):
        p = requests.get((wiki_images+jpeg+png+webp)[i])
        out = open(f'{folder+str(1000+i)}.png', 'wb')
        out.write(p.content)
        out.close()

    for i in os.listdir(folder):
        if imghdr.what(folder+i) != 'png' and imghdr.what(folder+i) != 'jpeg' and imghdr.what(folder+i) != 'webp':
            os.remove(folder+i)

    # return(str(time.time() - start_time) + ' | ' + str(answer_num[max_i][0]) + ' | ' + str(int(answer_num[max_i][1]*100)) + '% | ' + str(answer_num[max_i][2]))
    return(' | ' + str(time.time() - start_time) + ' | \n' + wiki_text + '\n\n' + ('\n').join(answers[:10]))

# print(search_('мифы о драконах'))
# reborn()
# db.sqlConnect('bot_db')
# db.sqlDel('dict', 'id > 84')
# db.sqlDisconnect()